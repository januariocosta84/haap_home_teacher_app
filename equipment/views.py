import io
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse_lazy
from django.views.generic import CreateView, ListView, UpdateView, DeleteView, DetailView
from django.utils import timezone
from django.db.models import Q, Count

from klase.models import Classroom
from preschools.models import Preschool
from core.models import User

from .forms import EquipmentForm, EquipmentAssignmentForm
from .models import Equipment, EquipmentAssignmentHistory, EquipmentType
from core.audit import log_action


class AdminOnlyMixin(UserPassesTestMixin):
    """Mixin to restrict access to admin users"""
    def test_func(self):
        return self.request.user.role == 'moe_admin'


class EquipmentCreateView(LoginRequiredMixin, AdminOnlyMixin, CreateView):

    model = Equipment
    form_class = EquipmentForm
    template_name = 'equipment/equipment_form.html'
    success_url = reverse_lazy('equipment:equipment_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['equipment_types'] = EquipmentType.objects.all()
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        obj = form.instance
        messages.success(self.request, 'Ekipamentu konsege rejistu ho susesu.')
        log_action(
            request=self.request, action='create', module='Ekipamentu',
            description=f"Kria ekipamentu: {obj.equipment_type.name} — {obj.model_number}",
            record_id=str(obj.pk), record_name=str(obj),
            new_value={'type': obj.equipment_type.name, 'model': obj.model_number,
                       'serial': obj.serial_number or '', 'status': obj.status},
        )
        return response

    def form_invalid(self, form):
        messages.error(self.request, form.errors)
        return super().form_invalid(form)


class EquipmentListView(LoginRequiredMixin, AdminOnlyMixin, ListView):

    model = Equipment
    template_name = 'equipment/equipment_list.html'
    context_object_name = 'equipments'
    paginate_by = 20

    def get_queryset(self):
        queryset = Equipment.objects.select_related(
            'equipment_type',
            'preschool',
            'classroom',
            'teacher'
        ).order_by('-created_at')

        # Filter by equipment type (FK id)
        equipment_type = self.request.GET.get('equipment_type')
        if equipment_type:
            queryset = queryset.filter(equipment_type_id=equipment_type)

        # Filter by status
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)

        # Filter by preschool
        preschool = self.request.GET.get('preschool')
        if preschool:
            queryset = queryset.filter(preschool_id=preschool)

        # Search by serial number or model
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(serial_number__icontains=search) |
                Q(model_number__icontains=search)
            )

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['equipment_types'] = EquipmentType.objects.all()
        context['statuses'] = Equipment.STATUS_CHOICES
        context['preschools'] = Preschool.objects.all()
        query = self.request.GET.copy()
        query.pop('page', None)
        context['querystring'] = query.urlencode()
        return context


class EquipmentDetailView(LoginRequiredMixin, AdminOnlyMixin, DetailView):

    model = Equipment
    template_name = 'equipment/equipment_detail.html'
    context_object_name = 'equipment'
    pk_url_kwarg = 'pk'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['assignment_history'] = (
            self.object.assignment_history.all().order_by('-changed_at')
        )
        return context


class EquipmentUpdateView(LoginRequiredMixin, AdminOnlyMixin, UpdateView):

    model = Equipment
    form_class = EquipmentForm
    template_name = 'equipment/equipment_form.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['equipment_types'] = EquipmentType.objects.all()
        return context

    def get_success_url(self):
        return reverse_lazy('equipment:equipment_detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        # Check if assignment changed
        if form.has_changed():
            old_preschool = self.object.preschool
            old_classroom = self.object.classroom
            old_teacher = self.object.teacher

            response = super().form_valid(form)

            # Log the assignment change if relevant fields changed
            if (old_preschool != form.cleaned_data.get('preschool') or
                old_classroom != form.cleaned_data.get('classroom') or
                old_teacher != form.cleaned_data.get('teacher')):

                EquipmentAssignmentHistory.objects.create(
                    equipment=self.object,
                    old_preschool=old_preschool,
                    old_classroom=old_classroom,
                    old_teacher=old_teacher,
                    new_preschool=form.cleaned_data.get('preschool'),
                    new_classroom=form.cleaned_data.get('classroom'),
                    new_teacher=form.cleaned_data.get('teacher'),
                    changed_by=self.request.user,
                    change_reason='Direct equipment update'
                )

            messages.success(self.request, 'Ekipamentu updates ho susesu.')
            log_action(
                request=self.request, action='update', module='Ekipamentu',
                description=f"Atualiza ekipamentu: {self.object}",
                record_id=str(self.object.pk), record_name=str(self.object),
                new_value={'status': self.object.status, 'model': self.object.model_number,
                           'preschool': str(self.object.preschool or ''), 'classroom': str(self.object.classroom or '')},
            )
            return response
        else:
            return super().form_valid(form)


class EquipmentAssignmentChangeView(LoginRequiredMixin, AdminOnlyMixin, UpdateView):

    model = Equipment
    form_class = EquipmentAssignmentForm
    template_name = 'equipment/equipment_assignment_form.html'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs.pop('instance', None)  # Remove instance for the non-model form
        return kwargs

    def form_valid(self, form):
        action = form.cleaned_data.get('action')
        # self.object is the Equipment instance fetched by get_object() before
        # form_valid runs, so it still holds the original assignment values here.
        old_equipment = self.object

        try:
            if action == 'reassign':
                new_preschool = form.cleaned_data.get('preschool')
                new_classroom = form.cleaned_data.get('classroom')
                new_teacher = form.cleaned_data.get('teacher')

                # Log the assignment change
                EquipmentAssignmentHistory.objects.create(
                    equipment=self.object,
                    old_preschool=old_equipment.preschool,
                    old_classroom=old_equipment.classroom,
                    old_teacher=old_equipment.teacher,
                    new_preschool=new_preschool,
                    new_classroom=new_classroom,
                    new_teacher=new_teacher,
                    changed_by=self.request.user,
                    change_reason=form.cleaned_data.get('change_reason')
                )

                # Update equipment
                self.object.preschool = new_preschool
                self.object.classroom = new_classroom
                self.object.teacher = new_teacher
                self.object.save()

                messages.success(self.request, 'Atribisaun ekipamentu updates ho susesu.')
                log_action(
                    request=self.request, action='update', module='Ekipamentu',
                    description=f"Muda atribuisaun: {self.object} → {new_preschool or new_teacher or '—'}",
                    record_id=str(self.object.pk), record_name=str(self.object),
                    previous_value={'preschool': str(old_equipment.preschool or ''), 'classroom': str(old_equipment.classroom or '')},
                    new_value={'preschool': str(new_preschool or ''), 'classroom': str(new_classroom or '')},
                )

            elif action == 'delete':
                # Log the deletion of assignment
                EquipmentAssignmentHistory.objects.create(
                    equipment=self.object,
                    old_preschool=old_equipment.preschool,
                    old_classroom=old_equipment.classroom,
                    old_teacher=old_equipment.teacher,
                    changed_by=self.request.user,
                    change_reason=form.cleaned_data.get('change_reason') or 'Assignment deleted'
                )

                # Clear assignment
                self.object.preschool = None
                self.object.classroom = None
                self.object.teacher = None
                self.object.save()

                messages.success(self.request, 'Atribisaun ekipamentu removidu ho susesu.')
                log_action(
                    request=self.request, action='update', module='Ekipamentu',
                    description=f"Hasai atribuisaun ekipamentu: {self.object}",
                    record_id=str(self.object.pk), record_name=str(self.object),
                )

            elif action == 'retire':
                # Log the retirement
                EquipmentAssignmentHistory.objects.create(
                    equipment=self.object,
                    old_preschool=old_equipment.preschool,
                    old_classroom=old_equipment.classroom,
                    old_teacher=old_equipment.teacher,
                    changed_by=self.request.user,
                    change_reason=form.cleaned_data.get('change_reason') or 'Equipment retired'
                )

                # Mark as retired
                self.object.status = 'retired'
                self.object.preschool = None
                self.object.classroom = None
                self.object.teacher = None
                self.object.save()

                messages.success(self.request, 'Ekipamentu retiradu ho susesu.')
                log_action(
                    request=self.request, action='deactivate', module='Ekipamentu',
                    description=f"Retira ekipamentu: {self.object}",
                    record_id=str(self.object.pk), record_name=str(self.object),
                )

            return redirect(self.get_success_url())

        except Exception as e:
            messages.error(self.request, f'Erro: {str(e)}')
            return redirect(self.get_success_url())

    def get_success_url(self):
        return reverse_lazy('equipment:equipment_detail', kwargs={'pk': self.object.pk})


class EquipmentDeleteView(LoginRequiredMixin, AdminOnlyMixin, DeleteView):

    model = Equipment
    template_name = 'equipment/equipment_confirm_delete.html'
    success_url = reverse_lazy('equipment:equipment_list')

    def form_valid(self, form):
        obj = self.get_object()
        record_id, record_name = str(obj.pk), str(obj)
        response = super().form_valid(form)
        messages.success(self.request, 'Ekipamentu deleta ho susesu.')
        log_action(
            request=self.request, action='delete', module='Ekipamentu',
            description=f"Apaga ekipamentu: {record_name}",
            record_id=record_id, record_name=record_name,
        )
        return response


class EquipmentByPreschoolView(LoginRequiredMixin, AdminOnlyMixin, ListView):

    model = Equipment
    template_name = 'equipment/equipment_by_preschool.html'
    context_object_name = 'equipments'
    paginate_by = 20

    def get_queryset(self):
        preschool_id = self.kwargs.get('preschool_id')
        return Equipment.objects.filter(
            preschool_id=preschool_id
        ).select_related('preschool', 'classroom', 'teacher')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        preschool_id = self.kwargs.get('preschool_id')
        context['preschool'] = get_object_or_404(Preschool, pk=preschool_id)
        return context


class EquipmentByClassroomView(LoginRequiredMixin, AdminOnlyMixin, ListView):

    model = Equipment
    template_name = 'equipment/equipment_by_classroom.html'
    context_object_name = 'equipments'
    paginate_by = 20

    def get_queryset(self):
        classroom_id = self.kwargs.get('classroom_id')
        return Equipment.objects.filter(
            classroom_id=classroom_id
        ).select_related('preschool', 'classroom', 'teacher')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        classroom_id = self.kwargs.get('classroom_id')
        context['classroom'] = get_object_or_404(Classroom, pk=classroom_id)
        return context


class EquipmentByTeacherView(LoginRequiredMixin, AdminOnlyMixin, ListView):

    model = Equipment
    template_name = 'equipment/equipment_by_teacher.html'
    context_object_name = 'equipments'
    paginate_by = 20

    def get_queryset(self):
        teacher_id = self.kwargs.get('teacher_id')
        return Equipment.objects.filter(
            teacher_id=teacher_id
        ).select_related('preschool', 'classroom', 'teacher')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        teacher_id = self.kwargs.get('teacher_id')
        context['teacher'] = get_object_or_404(User, pk=teacher_id, role='teacher')
        return context


# AJAX VIEW

def load_classrooms(request):

    preschool_id = request.GET.get('preschool_id')

    classrooms = Classroom.objects.filter(
        preschool_id=preschool_id
    ).values('id', 'name')

    return JsonResponse(list(classrooms), safe=False)


# ── Equipment Type Management ─────────────────────────────────────────────────

class EquipmentTypeListView(LoginRequiredMixin, AdminOnlyMixin, ListView):
    model = EquipmentType
    template_name = 'equipment/equipment_type_list.html'
    context_object_name = 'equipment_types'


EQUIPMENT_ICON_CHOICES = [
    'bi-tablet', 'bi-tablet-landscape', 'bi-laptop', 'bi-laptop-fill',
    'bi-projector', 'bi-projector-fill', 'bi-display', 'bi-display-fill',
    'bi-tv', 'bi-tv-fill', 'bi-phone', 'bi-phone-fill',
    'bi-usb-plug', 'bi-usb-plug-fill', 'bi-plug', 'bi-plug-fill',
    'bi-hdd', 'bi-hdd-fill', 'bi-cpu', 'bi-cpu-fill',
    'bi-server', 'bi-router', 'bi-router-fill', 'bi-wifi',
    'bi-camera', 'bi-camera-fill', 'bi-printer', 'bi-printer-fill',
    'bi-keyboard', 'bi-mouse', 'bi-speaker', 'bi-headphones',
    'bi-battery-full', 'bi-lightning-charge', 'bi-tools', 'bi-box-seam',
]


class EquipmentTypeCreateView(LoginRequiredMixin, AdminOnlyMixin, CreateView):
    model = EquipmentType
    fields = ['name', 'icon', 'serial_number_required']
    template_name = 'equipment/equipment_type_form.html'
    success_url = reverse_lazy('equipment:equipment_type_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['icon_choices'] = EQUIPMENT_ICON_CHOICES
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        obj = form.instance
        messages.success(self.request, f'Tipu "{obj.name}" adisionadu ho susesu.')
        log_action(
            request=self.request, action='create', module='Tipu Ekipamentu',
            description=f"Kria tipu ekipamentu: {obj.name}",
            record_id=str(obj.pk), record_name=obj.name,
            new_value={'name': obj.name, 'serial_required': obj.serial_number_required},
        )
        return response


class EquipmentTypeUpdateView(LoginRequiredMixin, AdminOnlyMixin, UpdateView):
    model = EquipmentType
    fields = ['name', 'icon', 'serial_number_required']
    template_name = 'equipment/equipment_type_form.html'
    success_url = reverse_lazy('equipment:equipment_type_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['icon_choices'] = EQUIPMENT_ICON_CHOICES
        return context

    def form_valid(self, form):
        old_name = self.object.name
        response = super().form_valid(form)
        obj = form.instance
        messages.success(self.request, f'Tipu "{obj.name}" atualizado ho susesu.')
        log_action(
            request=self.request, action='update', module='Tipu Ekipamentu',
            description=f"Atualiza tipu ekipamentu: {obj.name}",
            record_id=str(obj.pk), record_name=obj.name,
            previous_value={'name': old_name},
            new_value={'name': obj.name, 'serial_required': obj.serial_number_required},
        )
        return response


class EquipmentTypeDeleteView(LoginRequiredMixin, AdminOnlyMixin, DeleteView):
    model = EquipmentType
    template_name = 'equipment/equipment_type_confirm_delete.html'
    success_url = reverse_lazy('equipment:equipment_type_list')

    def form_valid(self, form):
        obj = self.get_object()
        record_id, record_name = str(obj.pk), obj.name
        try:
            response = super().form_valid(form)
            messages.success(self.request, 'Tipu ekipamentu deletadu ho susesu.')
            log_action(
                request=self.request, action='delete', module='Tipu Ekipamentu',
                description=f"Apaga tipu ekipamentu: {record_name}",
                record_id=record_id, record_name=record_name,
            )
            return response
        except Exception:
            messages.error(self.request, 'La bele apaga tipu ne\'e tanba iha ekipamentu ho tipu ne\'e.')
            return redirect('equipment:equipment_type_list')


def equipment_type_info(request):
    """Return serial_number_required flag for a given EquipmentType id."""
    type_id = request.GET.get('id')
    try:
        et = EquipmentType.objects.get(pk=type_id)
        return JsonResponse({'serial_number_required': et.serial_number_required})
    except EquipmentType.DoesNotExist:
        return JsonResponse({'serial_number_required': False})


# ── Preschool Equipment Exports ───────────────────────────────────────────────

def _preschool_equipment_qs(preschool_id):
    return (
        Equipment.objects
        .filter(preschool_id=preschool_id)
        .select_related('equipment_type', 'preschool', 'classroom', 'teacher')
        .order_by('equipment_type__name', 'serial_number')
    )


@login_required
def export_preschool_equipment_excel(request, preschool_id):
    """Download all equipment for a preschool as an Excel workbook."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    preschool = get_object_or_404(Preschool, pk=preschool_id)
    qs = _preschool_equipment_qs(preschool_id)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = 'Sumáriu'

    hdr_fill  = PatternFill('solid', fgColor='1D4ED8')
    hdr_font  = Font(color='FFFFFF', bold=True, size=11)
    sub_fill  = PatternFill('solid', fgColor='EFF6FF')
    sub_font  = Font(color='1E3A8A', bold=True)
    thin      = Side(style='thin', color='CBD5E1')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    center    = Alignment(horizontal='center', vertical='center')

    # Title
    ws_sum.merge_cells('A1:D1')
    title_cell = ws_sum['A1']
    title_cell.value = f'Relatóriu Ekipamentu — {preschool.name}'
    title_cell.font  = Font(bold=True, size=14, color='0F172A')
    title_cell.alignment = center

    ws_sum.merge_cells('A2:D2')
    date_cell = ws_sum['A2']
    date_cell.value = f'Data: {timezone.localdate().strftime("%d %B %Y")}  |  Munisípiu: {preschool.municipality or "—"}'
    date_cell.font  = Font(size=10, color='64748B')
    date_cell.alignment = center

    ws_sum.append([])  # blank row

    # Header row
    headers = ['Tipu Ekipamentu', 'Total', 'Ativu', 'Estragu / Outros']
    ws_sum.append(headers)
    for col_idx, h in enumerate(headers, 1):
        cell = ws_sum.cell(row=4, column=col_idx)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = center
        cell.border    = border

    # Group by type
    from collections import defaultdict
    by_type = defaultdict(lambda: {'total': 0, 'active': 0, 'other': 0})
    for eq in qs:
        by_type[eq.equipment_type.name]['total'] += 1
        if eq.status == 'active':
            by_type[eq.equipment_type.name]['active'] += 1
        else:
            by_type[eq.equipment_type.name]['other'] += 1

    row_num = 5
    for type_name, counts in sorted(by_type.items()):
        ws_sum.append([type_name, counts['total'], counts['active'], counts['other']])
        for col_idx in range(1, 5):
            cell = ws_sum.cell(row=row_num, column=col_idx)
            cell.border = border
            if col_idx > 1:
                cell.alignment = center
        row_num += 1

    # Total row
    ws_sum.append(['TOTAL', sum(v['total'] for v in by_type.values()),
                   sum(v['active'] for v in by_type.values()),
                   sum(v['other'] for v in by_type.values())])
    for col_idx in range(1, 5):
        cell = ws_sum.cell(row=row_num, column=col_idx)
        cell.fill   = sub_fill
        cell.font   = sub_font
        cell.border = border
        cell.alignment = center

    ws_sum.column_dimensions['A'].width = 28
    for col in ['B', 'C', 'D']:
        ws_sum.column_dimensions[col].width = 14
    ws_sum.row_dimensions[1].height = 28
    ws_sum.row_dimensions[4].height = 22

    # ── Sheet 2: Full list ───────────────────────────────────────────────────
    ws_list = wb.create_sheet('Lista Kompletu')

    list_headers = ['#', 'Tipu', 'Modelu', 'Numeru Série', 'Status', 'Klase', 'Nóta']
    ws_list.append(list_headers)
    for col_idx, h in enumerate(list_headers, 1):
        cell = ws_list.cell(row=1, column=col_idx)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = center
        cell.border    = border

    STATUS_TL = {'active': 'Ativu', 'inactive': 'Inativu', 'damaged': 'Estragu', 'retired': 'Retirado'}
    for idx, eq in enumerate(qs, 1):
        row = [
            idx,
            eq.equipment_type.name,
            eq.model_number or '—',
            eq.serial_number or '—',
            STATUS_TL.get(eq.status, eq.status),
            eq.classroom.name if eq.classroom else '—',
            eq.notes or '',
        ]
        ws_list.append(row)
        for col_idx, val in enumerate(row, 1):
            cell = ws_list.cell(row=idx + 1, column=col_idx)
            cell.border = border
            if col_idx == 1:
                cell.alignment = center

    col_widths = [6, 18, 18, 20, 12, 18, 30]
    for col_idx, width in enumerate(col_widths, 1):
        ws_list.column_dimensions[get_column_letter(col_idx)].width = width
    ws_list.row_dimensions[1].height = 22

    # Output
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'ekipamentu_{preschool.name.replace(" ", "_")}_{timezone.localdate().isoformat()}.xlsx'
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def export_preschool_equipment_pdf(request, preschool_id):
    """Download all equipment for a preschool as a PDF."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    preschool = get_object_or_404(Preschool, pk=preschool_id)
    qs = list(_preschool_equipment_qs(preschool_id))

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=15*mm, bottomMargin=15*mm)

    styles  = getSampleStyleSheet()
    BLUE    = colors.HexColor('#1D4ED8')
    LT_BLUE = colors.HexColor('#EFF6FF')
    GREY    = colors.HexColor('#64748B')
    D_GREY  = colors.HexColor('#0F172A')

    title_style = ParagraphStyle('title', parent=styles['Heading1'],
                                 fontSize=16, textColor=D_GREY, spaceAfter=2*mm)
    sub_style   = ParagraphStyle('sub', parent=styles['Normal'],
                                 fontSize=10, textColor=GREY, spaceAfter=6*mm)

    elements = [
        Paragraph(f'Relatóriu Ekipamentu — {preschool.name}', title_style),
        Paragraph(
            f'Munisípiu: {preschool.municipality or "—"} &nbsp;&nbsp;|&nbsp;&nbsp; '
            f'Data: {timezone.localdate().strftime("%d %B %Y")} &nbsp;&nbsp;|&nbsp;&nbsp; '
            f'Total: {len(qs)} ekipamentu',
            sub_style),
    ]

    # Summary table
    from collections import defaultdict
    by_type = defaultdict(lambda: {'total': 0, 'active': 0, 'other': 0})
    for eq in qs:
        by_type[eq.equipment_type.name]['total'] += 1
        if eq.status == 'active':
            by_type[eq.equipment_type.name]['active'] += 1
        else:
            by_type[eq.equipment_type.name]['other'] += 1

    sum_data = [['Tipu Ekipamentu', 'Total', 'Ativu', 'Estragu / Outros']]
    for type_name, counts in sorted(by_type.items()):
        sum_data.append([type_name, str(counts['total']), str(counts['active']), str(counts['other'])])
    sum_data.append(['TOTAL',
                     str(sum(v['total'] for v in by_type.values())),
                     str(sum(v['active'] for v in by_type.values())),
                     str(sum(v['other'] for v in by_type.values()))])

    sum_table = Table(sum_data, colWidths=[80*mm, 30*mm, 30*mm, 50*mm])
    sum_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), BLUE),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 9),
        ('ALIGN',      (1, 0), (-1, -1), 'CENTER'),
        ('ALIGN',      (0, 0), (0, -1), 'LEFT'),
        ('BACKGROUND', (0, -1), (-1, -1), LT_BLUE),
        ('FONTNAME',   (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F8FAFC')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('TOPPADDING',    (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING',   (0, 0), (-1, -1), 6),
    ]))
    elements += [sum_table, Spacer(1, 8*mm)]

    # Detail table
    STATUS_TL = {'active': 'Ativu', 'inactive': 'Inativu', 'damaged': 'Estragu', 'retired': 'Retirado'}
    detail_data = [['#', 'Tipu', 'Modelu', 'Numeru Série', 'Status', 'Klase']]
    for idx, eq in enumerate(qs, 1):
        detail_data.append([
            str(idx),
            eq.equipment_type.name,
            eq.model_number or '—',
            eq.serial_number or '—',
            STATUS_TL.get(eq.status, eq.status),
            eq.classroom.name if eq.classroom else '—',
        ])

    detail_table = Table(detail_data, colWidths=[12*mm, 40*mm, 40*mm, 50*mm, 25*mm, 40*mm])
    detail_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), BLUE),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 8),
        ('ALIGN',      (0, 0), (0, -1), 'CENTER'),
        ('ALIGN',      (4, 0), (4, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#E2E8F0')),
        ('TOPPADDING',    (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING',   (0, 0), (-1, -1), 5),
    ]))
    elements.append(detail_table)

    doc.build(elements)
    buffer.seek(0)
    filename = f'ekipamentu_{preschool.name.replace(" ", "_")}_{timezone.localdate().isoformat()}.pdf'
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
