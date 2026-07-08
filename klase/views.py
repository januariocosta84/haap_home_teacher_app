import csv
from io import BytesIO

from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib import messages
from django.views import View
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator

from core.models import Child, User
from .models import Classroom, ClassroomChild
from django.views.generic import ListView
from .forms import AddChildToClassForm, ChildCodeEnrollmentForm
from django.contrib.auth.mixins import LoginRequiredMixin
from django.shortcuts import get_object_or_404
from django.views.generic import TemplateView
from django.utils.text import slugify
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

from preschools.models import PreschoolTeacher
from preschools.models import Preschool
from .models import Classroom


@method_decorator(login_required, name='dispatch')
class TeacherSchoolListView(View):
    """Teacher's dashboard showing all approved schools with assigned classes."""
    template_name = "klase/teacher_school_list.html"

    def get(self, request):
        # Get approved teacher-school relations
        preschool_teachers = PreschoolTeacher.objects.filter(
            teacher=request.user,
            is_active=True,
            is_approved=True
        ).select_related('preschool')

        preschools_data = []
        total_students = 0

        for pt in preschool_teachers:
            preschool = pt.preschool
            # Get classrooms in this school taught by this teacher
            classrooms = Classroom.objects.filter(
                preschool=preschool,
                teacher=request.user
            ).prefetch_related('enrollments')
            
            classroom_count = classrooms.count()
            if classroom_count == 0:
                continue

            student_count = sum(c.enrollments.filter(is_active=True).count() for c in classrooms)
            total_students += student_count

            preschools_data.append({
                'id': preschool.id,
                'name': preschool.name,
                'preschool_type': preschool.get_preschool_type_display(),
                'classroom_count': classroom_count,
                'student_count': student_count,
                'is_primary': pt.is_primary,
            })

        context = {
            'preschools': preschools_data,
            'schools': preschools_data,  # For sidebar
            'total_schools': len(preschools_data),
            'total_students': total_students,
        }

        return render(request, self.template_name, context)


@method_decorator(login_required, name='dispatch')
class SchoolClassroomListView(View):
    """Show all classrooms in a specific school"""
    template_name = "klase/school_classroom_list.html"

    def get(self, request, preschool_id):
        # Verify teacher has access to this preschool
        preschool = get_object_or_404(
            Preschool,
            id=preschool_id
        )

        preschool_teacher = get_object_or_404(
            PreschoolTeacher,
            preschool=preschool,
            teacher=request.user,
            is_active=True
        )

        # Get classrooms in this school taught by this teacher
        classrooms = Classroom.objects.filter(
            preschool=preschool,
            teacher=request.user
        ).prefetch_related('enrollments', 'enrollments__child')

        classroom_data = []
        total_students = 0
        group_a_count = 0
        group_b_count = 0

        for classroom in classrooms:
            enrollments = classroom.enrollments.filter(is_active=True)
            student_count = enrollments.count()
            total_students += student_count

            # Count by age group
            for enrollment in enrollments:
                if enrollment.child.age_group == 'A':
                    group_a_count += 1
                elif enrollment.child.age_group == 'B':
                    group_b_count += 1

            classroom_data.append({
                'id': classroom.id,
                'name': classroom.name,
                'group': classroom.group,
                'student_count': student_count,
                'students': list(enrollments.values('child__id', 'child__first_name', 'child__user_id', 'child__age_group')),
            })

        context = {
            'preschool': {
                'id': preschool.id,
                'name': preschool.name,
                'preschool_type': preschool.get_preschool_type_display(),
            },
            'classrooms': classroom_data,
            'total_classrooms': len(classroom_data),
            'total_students': total_students,
            'group_a_count': group_a_count,
            'group_b_count': group_b_count,
            # For sidebar
            'schools': [],
            'total_schools': 1,
        }

        # Get all schools for sidebar
        all_preschools = PreschoolTeacher.objects.filter(
            teacher=request.user,
            is_active=True
        ).select_related('preschool')

        schools_for_sidebar = []
        for pt in all_preschools:
            ps = pt.preschool
            classrooms_count = Classroom.objects.filter(
                preschool=ps,
                teacher=request.user
            ).count()
            schools_for_sidebar.append({
                'id': ps.id,
                'name': ps.name,
                'classroom_count': classrooms_count,
            })

        context['schools'] = schools_for_sidebar
        context['total_schools'] = len(schools_for_sidebar)

        return render(request, self.template_name, context)


@method_decorator(login_required, name='dispatch')
class ClassroomDetailView(View):
    """Show classroom details with enrolled students and option to add children"""
    template_name = "klase/classroom_detail.html"

    def get_classroom_queryset(self, request):
        queryset = Classroom.objects.select_related('preschool', 'teacher')
        if request.user.role == 'moe_admin':
            return queryset
        return queryset.filter(teacher=request.user)

    def get(self, request, classroom_id):
        classroom = get_object_or_404(
            self.get_classroom_queryset(request),
            id=classroom_id
        )

        enrollments = classroom.enrollments.filter(is_active=True).select_related('child', 'child__parent')
        form = ChildCodeEnrollmentForm()

        # Get all schools for sidebar
        schools_for_sidebar = []
        total_students_all = 0
        if request.user.role == 'moe_admin':
            classrooms_list = Classroom.objects.filter(
                preschool=classroom.preschool
            ).prefetch_related('enrollments')
            students_count = sum(c.enrollments.filter(is_active=True).count() for c in classrooms_list)
            total_students_all += students_count
            schools_for_sidebar.append({
                'id': classroom.preschool.id,
                'name': classroom.preschool.name,
                'classroom_count': classrooms_list.count(),
            })
        else:
            all_preschools = PreschoolTeacher.objects.filter(
                teacher=request.user,
                is_active=True
            ).select_related('preschool')

            for pt in all_preschools:
                ps = pt.preschool
                classrooms_list = Classroom.objects.filter(
                    preschool=ps,
                    teacher=request.user
                ).prefetch_related('enrollments')
                classrooms_count = classrooms_list.count()
                students_count = sum(c.enrollments.filter(is_active=True).count() for c in classrooms_list)
                total_students_all += students_count
                schools_for_sidebar.append({
                    'id': ps.id,
                    'name': ps.name,
                    'classroom_count': classrooms_count,
                })

        context = {
            'classroom': {
                'id': classroom.id,
                'name': classroom.name,
                'group': classroom.group,
                'preschool_id': classroom.preschool.id,
                'preschool_name': classroom.preschool.name,
            },
            'students': list(enrollments.values(
                'id',
                'child__id',
                'child__first_name',
                'child__user_id',
                'child__age_group',
                'child__parent__first_name',
                'child__parent__last_name',
                'child__parent__username',
                'child__parent__whatsapp_number',
            )),
            'student_count': enrollments.count(),
            'form': form,
            # For sidebar
            'schools': schools_for_sidebar,
            'total_schools': len(schools_for_sidebar),
            'total_students': total_students_all,
        }

        return render(request, self.template_name, context)

    def post(self, request, classroom_id):
        import datetime

        if request.user.role == 'moe_admin':
            messages.error(request, "MoE admin bele haree klase, maibe labele aumenta alunu iha pajina ida ne'e.")
            return redirect('klase:classroom_detail', classroom_id=classroom_id)

        classroom = get_object_or_404(
            Classroom,
            id=classroom_id,
            teacher=request.user
        )

        form = ChildCodeEnrollmentForm(request.POST)

        if not form.is_valid():
            messages.error(request, "Kodigu invalidu.")
            return redirect('klase:classroom_detail', classroom_id=classroom.id)

        child_code = form.cleaned_data['child_code']

        # 1. Find child by user_id (the code)
        child = Child.objects.filter(user_id=child_code).first()

        if not child:
            messages.error(
                request,
                f"Kodigu '{child_code}' la existe iha sistema. Favor verifika kodigu ho los."
            )
            return redirect('klase:classroom_detail', classroom_id=classroom.id)

        # 2. Validate registered age group matches classroom group
        if child.age_group != classroom.group:
            group_label = {'A': 'Grupo A (Tinan 3–4)', 'B': 'Grupo B (Tinan 5–6)'}
            messages.error(
                request,
                f"Labarik '{child.first_name}' pertense {group_label.get(child.age_group, child.age_group)}, "
                f"maibé klase ida ne'e uza {group_label.get(classroom.group, classroom.group)}. "
                f"Labarik labele rejista iha grupo ne'ebé la tuir."
            )
            return redirect('klase:classroom_detail', classroom_id=classroom.id)

        # 3. Validate actual age from year_of_birth
        current_year = datetime.date.today().year
        child_age = current_year - child.year_of_birth
        age_ranges = {'A': (3, 4), 'B': (5, 6)}
        min_age, max_age = age_ranges[classroom.group]

        if not (min_age <= child_age <= max_age):
            messages.error(
                request,
                f"Labarik '{child.first_name}' iha tinan {child_age} "
                f"({child.year_of_birth}), maibé klase {classroom.group} presiza labarik tinan {min_age}–{max_age}. "
                f"Labarik la elegível atu rejista iha klase ida ne'e."
            )
            return redirect('klase:classroom_detail', classroom_id=classroom.id)

        # 4. Check if already actively enrolled in THIS classroom
        this_enrollment = ClassroomChild.objects.filter(
            classroom=classroom, child=child
        ).first()

        if this_enrollment and this_enrollment.is_active:
            messages.warning(
                request,
                f"{child.first_name} rejistadu ona iha klase '{classroom.name}'."
            )
            return redirect('klase:classroom_detail', classroom_id=classroom.id)

        # 5. Check if actively enrolled in ANY OTHER classroom (one-class rule)
        other_enrollment = ClassroomChild.objects.filter(
            child=child,
            is_active=True
        ).exclude(classroom=classroom).select_related(
            'classroom', 'classroom__preschool'
        ).first()

        if other_enrollment:
            messages.error(
                request,
                f"{child.first_name} rejistadu ona iha klase "
                f"'{other_enrollment.classroom.name}' "
                f"({other_enrollment.classroom.preschool.name}). "
                f"Favor hasai labarik hosi klase ne'ebé atual molok rejista iha klase seluk."
            )
            return redirect('klase:classroom_detail', classroom_id=classroom.id)

        # 6. All checks passed — enroll
        if this_enrollment:
            # reactivate a previously removed enrollment in this same class
            this_enrollment.is_active = True
            this_enrollment.save(update_fields=['is_active'])
        else:
            ClassroomChild.objects.create(
                classroom=classroom,
                child=child,
                is_active=True
            )

        messages.success(
            request,
            f"{child.first_name} adisiona ona ba klase '{classroom.name}' ho susesu."
        )
        return redirect('klase:classroom_detail', classroom_id=classroom.id)


@method_decorator(login_required, name='dispatch')
class RemoveStudentFromClassroomView(View):
    """Allow a teacher to remove a child from one of their classrooms."""

    def post(self, request, classroom_id, enrollment_id):
        classroom = get_object_or_404(
            Classroom,
            id=classroom_id,
            teacher=request.user
        )
        enrollment = get_object_or_404(
            ClassroomChild.objects.select_related('child'),
            id=enrollment_id,
            classroom=classroom,
            is_active=True
        )

        child_name = enrollment.child.first_name
        enrollment.is_active = False
        enrollment.save(update_fields=['is_active'])

        messages.success(request, f"{child_name} remove ona hosi klase {classroom.name}.")
        return redirect('klase:classroom_detail', classroom_id=classroom.id)


@method_decorator(login_required, name='dispatch')
class DownloadClassroomChildrenView(View):
    """Download active children in a teacher-owned classroom."""

    headers = [
        'No',
        'Child name',
        'Student code',
        'Age group',
        'Year of birth',
        'Parent name',
        'Parent WhatsApp',
        'Enrolled at',
    ]

    def get(self, request, classroom_id):
        classrooms = Classroom.objects.select_related('preschool')
        if request.user.role != 'moe_admin':
            classrooms = classrooms.filter(teacher=request.user)
        classroom = get_object_or_404(classrooms, id=classroom_id)
        enrollments = (
            classroom.enrollments
            .filter(is_active=True)
            .select_related('child', 'child__parent')
            .order_by('child__first_name')
        )

        rows = self.get_rows(enrollments)
        filename = slugify(f"{classroom.name}-children") or "classroom-children"
        export_format = request.GET.get('format', 'xlsx').lower()

        if export_format == 'csv':
            return self.csv_response(filename, rows)
        if export_format == 'pdf':
            return self.pdf_response(filename, classroom, rows)
        return self.excel_response(filename, rows)

    def get_rows(self, enrollments):
        rows = []
        for index, enrollment in enumerate(enrollments, start=1):
            child = enrollment.child
            parent = child.parent
            rows.append([
                index,
                child.first_name,
                child.user_id,
                child.get_age_group_display(),
                child.year_of_birth,
                parent.get_full_name() or parent.username,
                parent.whatsapp_number,
                enrollment.enrolled_at.strftime('%Y-%m-%d %H:%M'),
            ])
        return rows

    def csv_response(self, filename, rows):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'

        writer = csv.writer(response)
        writer.writerow(self.headers)
        writer.writerows(rows)
        return response

    def excel_response(self, filename, rows):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Children'

        sheet.append(self.headers)
        for row in rows:
            sheet.append(row)

        header_fill = PatternFill(start_color='D9EAF7', end_color='D9EAF7', fill_type='solid')
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

        for column_cells in sheet.columns:
            width = max(len(str(cell.value or '')) for cell in column_cells) + 2
            sheet.column_dimensions[column_cells[0].column_letter].width = min(width, 35)

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        return response

    def pdf_response(self, filename, classroom, rows):
        buffer = BytesIO()
        document = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=0.35 * inch,
            leftMargin=0.35 * inch,
            topMargin=0.35 * inch,
            bottomMargin=0.35 * inch,
        )

        styles = getSampleStyleSheet()
        story = [
            Paragraph(f"{classroom.name} - Child List", styles['Title']),
            Paragraph(classroom.preschool.name, styles['Normal']),
            Spacer(1, 0.15 * inch),
        ]

        table = Table([self.headers] + rows, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D9EAF7')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
        ]))
        story.append(table)

        document.build(story)
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
        return response


class TeacherDashboardView(View):

    template_name = "klase/teacher_dashboard.html"

    def get(self, request):

        # Get all classrooms for this teacher
        classrooms = Classroom.objects.filter(
            teacher=request.user
        ).prefetch_related("enrollments", "enrollments__child")

        classroom_data = []

        total_students = 0

        for c in classrooms:

            student_count = c.enrollments.filter(is_active=True).count()
            total_students += student_count

            classroom_data.append({
                "id": c.id,
                "class_name": c.class_name,
                "group": c.group,
                "student_count": student_count
            })

        context = {
            "classrooms": classroom_data,
            "total_classes": classrooms.count(),
            "total_students": total_students
        }

        return render(request, self.template_name, context)
class TeacherProDashboardView(View):

    template_name = "klase/teacher_dashboard_pro.html"

    def get(self, request):

        classrooms = Classroom.objects.filter(
            teacher=request.user
        ).prefetch_related("enrollments")

        data = []

        total_students = 0
        group_a = 0
        group_b = 0

        for c in classrooms:

            count = c.enrollments.filter(is_active=True).count()
            total_students += count

            group_a += c.enrollments.filter(child__age_group='A').count()
            group_b += c.enrollments.filter(child__age_group='B').count()

            data.append({
                "id": c.id,
                "name": c.class_name,
                "group": c.group,
                "students": count
            })

        return render(request, self.template_name, {
            "classrooms": data,
            "total_classes": classrooms.count(),
            "total_students": total_students,
            "group_a": group_a,
            "group_b": group_b,
        })
    
class AjaxRegisterChild(View):
    def post(self, request):

        classroom_id = request.POST.get("classroom_id")
        parent_code = request.POST.get("user_id")
        first_name = request.POST.get("first_name")

        classroom = Classroom.objects.get(id=classroom_id)

        parent = User.objects.filter(username=parent_code).first()

        if not parent:
            return JsonResponse({"error": "Parent not found"}, status=400)

        child = Child.objects.create(
            parent=parent,
            first_name=first_name,
            year_of_birth=2019,
            age_group=classroom.group
        )

        ClassroomChild.objects.create(
            classroom=classroom,
            child=child,
            added_by=request.user
        )

        return JsonResponse({"success": True})

class ClassroomListView(ListView):
    model = Classroom
    template_name = "klase/classroom_list.html"
    context_object_name = "classrooms"

    def get_queryset(self):
        return Classroom.objects.all()
    
class AddChildToClassroomView(View):
    template_name = "klase/classroom_detail.html"

    def post(self, request, classroom_id):

        classroom = get_object_or_404(
            Classroom,
            id=classroom_id,
            teacher=request.user
        )

        form = AddChildToClassForm(request.POST)

        if not form.is_valid():
            messages.error(request, "Invalid data.")
            return redirect("klase:classroom_detail", classroom_id=classroom.id)

        user_id = form.cleaned_data["user_id"]
        first_name = form.cleaned_data["first_name"]

        # 1. Find child
        child = Child.objects.filter(
            user_id=user_id,
            first_name__iexact=first_name
        ).first()

        if not child:
            messages.error(request, "Child not found.")
            return redirect("klase:classroom_detail", classroom_id=classroom.id)

        # 2. Check if already assigned
        if hasattr(child, "classroom_enrollment"):
            messages.error(
                request,
                f"{child.first_name} already belongs to a class."
            )
            return redirect("klase:classroom_detail", classroom_id=classroom.id)

        # 3. Age group validation
        if child.age_group != classroom.group:
            messages.error(
                request,
                "Child age group does not match classroom."
            )
            return redirect("klase:classroom_detail", classroom_id=classroom.id)

        # 4. Create enrollment
        ClassroomChild.objects.create(
            classroom=classroom,
            child=child,
            added_by=request.user
        )

        messages.success(request, "Child added successfully.")

        return redirect("klase:classroom_detail", classroom_id=classroom.id)

class EnrollChildView(LoginRequiredMixin, View):
    def post(self, request, classroom_id):
        classroom = get_object_or_404(Classroom, id=classroom_id)
        form = ChildCodeEnrollmentForm(request.POST)

        if form.is_valid():
            code = form.cleaned_data['child_code']
            child = get_object_or_404(Child, child_code=code)

            # deactivate previous active classroom
            ClassroomChild.objects.filter(
                child=child,
                is_active=True
            ).update(is_active=False)

            # assign to new classroom
            ClassroomChild.objects.get_or_create(
                classroom=classroom,
                child=child,
                defaults={'is_active': True}
            )

            messages.success(request, 'Child enrolled successfully.')

        return redirect('classroom_detail', classroom_id=classroom.id)
