import uuid
from datetime import datetime, timedelta
from io import BytesIO

from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, render
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.views.generic import ListView
from django.core.exceptions import PermissionDenied
from django.template.loader import render_to_string

from core.models import ActivityResult, TeacherActivityLog, User, Child
from klase.models import ClassroomChild

def is_valid_uuid(value):
    try:
        uuid.UUID(str(value))
        return True
    except (ValueError, TypeError):
        return False


PAGE_SIZE = 50  # rows per page


@login_required
def AppUsageLogListView(request):
    user = request.user

    # ── Fix N+1: ActivityResult has two parent paths:
    #   1. .parent (direct FK, db_column='pid') — used by template for location fields
    #   2. .student.parent (via Child FK) — used for role-based filtering
    # Both chains need full select_related to avoid per-row queries.
    queryset = (
        ActivityResult.objects
        .select_related(
            "parent",
            "parent__municipality",
            "parent__administrative_post",
            "parent__suco",
            "parent__aldeia",
            "student",
            "student__parent",
            "student__parent__municipality",
            "student__parent__administrative_post",
            "student__parent__suco",
            "student__parent__aldeia",
        )
        .order_by("-created_at")
    )

    # ── Role-based access control ────────────────────────────────────────────
    if user.role == "parent":
        queryset = queryset.filter(student__parent=user)
    elif user.role == "municipality_analyst":
        queryset = queryset.filter(
            student__parent__municipality=user.municipality
        )
    elif user.role == "teacher":
        queryset = queryset.filter(
            student__classroom_history__classroom__teacher=user,
            student__classroom_history__is_active=True
        ).distinct()
    elif user.role == "moe_admin":
        pass
    else:
        queryset = ActivityResult.objects.none()

    # ── Skip rows with no student/parent ────────────────────────────────────
    queryset = queryset.filter(
        student__isnull=False,
        student__parent__isnull=False
    )

    # ── Server-side filters ──────────────────────────────────────────────────
    activity = request.GET.get("activity", "").strip()
    if activity:
        queryset = queryset.filter(activity_name__icontains=activity)

    search_q = request.GET.get("q", "").strip()
    if search_q:
        queryset = queryset.filter(
            Q(activity_name__icontains=search_q)
            | Q(student__first_name__icontains=search_q)
            | Q(student__parent__first_name__icontains=search_q)
            | Q(student__parent__last_name__icontains=search_q)
        )

    parent_id = request.GET.get("parent")
    if parent_id and is_valid_uuid(parent_id):
        queryset = queryset.filter(student__parent__id=parent_id)

    student_id = request.GET.get("student")
    if student_id and is_valid_uuid(student_id):
        queryset = queryset.filter(student__id=student_id)

    date_from = request.GET.get("date_from", "").strip()
    if date_from:
        try:
            queryset = queryset.filter(
                created_at__gte=datetime.strptime(date_from, "%Y-%m-%d")
            )
        except ValueError:
            date_from = ""

    date_to = request.GET.get("date_to", "").strip()
    if date_to:
        try:
            queryset = queryset.filter(
                created_at__lte=datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1)
            )
        except ValueError:
            date_to = ""

    # ── Aggregate stats from the full filtered queryset (before pagination) ─
    from django.db.models import Count
    stats = queryset.aggregate(
        total_activities=Count("id"),
        unique_students=Count("student", distinct=True),
        achieved=Count(
            "id",
            filter=Q(activity_result__isnull=False)
            & ~Q(activity_result="")
            & ~Q(activity_result="Tentadu"),
        ),
    )
    total_activities = stats["total_activities"]
    unique_students = stats["unique_students"]
    achieved = stats["achieved"]
    rate = round(achieved / total_activities * 100, 1) if total_activities else 0

    # ── Server-side pagination ───────────────────────────────────────────────
    paginator = Paginator(queryset, PAGE_SIZE)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # ── Split category1 for current page only (O(page_size) not O(total)) ──
    logs_with_split = []
    for log in page_obj:
        if log.category1 and "-" in log.category1:
            parts = log.category1.split("-", 1)
            log.tema = parts[0].strip()
            log.tipu = parts[1].strip() if len(parts) > 1 else ""
        else:
            log.tema = log.category1 or ""
            log.tipu = ""
        logs_with_split.append(log)

    context = {
        "logs": logs_with_split,
        "page_obj": page_obj,
        "paginator": paginator,
        "is_paginated": paginator.num_pages > 1,
        "total_activities": total_activities,
        "unique_students": unique_students,
        "total_achieved": achieved,
        "achievement_rate": rate,
        "search_q": search_q,
        "activity_filter": activity,
        "date_from": date_from,
        "date_to": date_to,
    }

    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        html = render_to_string("partials/logs_body.html", context, request=request)
        return JsonResponse({
            "html": html,
            "stats": {
                "total_activities": total_activities,
                "unique_students": unique_students,
                "total_achieved": achieved,
                "achievement_rate": rate,
            },
        })
    return render(request, "dashboards/logs.html", context)

class ChildActivityView(LoginRequiredMixin, ListView):
    model = ActivityResult
    template_name = "dashboards/child_activity.html"
    context_object_name = "activities"
    paginate_by = 10  # 👈 number of records per page

    def dispatch(self, request, *args, **kwargs):
        """
        Security check BEFORE loading the page
        """
        child_id = self.kwargs["child_id"]

        self.child = get_object_or_404(Child, id=child_id)

        is_parent = self.child.parent == request.user
        is_connected_teacher = (
            request.user.role == "teacher"
            and ClassroomChild.objects.filter(
                child=self.child,
                classroom__teacher=request.user,
                is_active=True
            ).exists()
        )
        is_moe_admin = request.user.role == "moe_admin"

        if not (is_parent or is_connected_teacher or is_moe_admin):
            raise PermissionDenied("Ita la iha permisaun atu haree labarik ida ne'e.")

        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        queryset = ActivityResult.objects.filter(
            student=self.child,
            parent=self.child.parent
        ).order_by("-created_at")

        # Split category1 into two attributes
        for obj in queryset:
            if obj.category1:
                if '-' in obj.category1:
                    parts = obj.category1.split('-')
                    obj.cat1_left = parts[0].strip()
                    obj.cat1_right = parts[1].strip() if len(parts) > 1 else ''
                else:
                    # Only one value, put it in left column
                    obj.cat1_left = obj.category1.strip()
                    obj.cat1_right = ''
            else:
                obj.cat1_left = ''
                obj.cat1_right = ''

            # Optional: split category2 and category3 similarly if needed
            obj.cat2_left = obj.category2 or ''
            obj.cat3_left = obj.category3 or ''

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["child"] = self.child

        from django.db.models import Count, Q
        stats = ActivityResult.objects.filter(
            student=self.child, parent=self.child.parent
        ).aggregate(
            total=Count("id"),
            achieved=Count("id", filter=Q(activity_result__isnull=False) & ~Q(activity_result="") & ~Q(activity_result="Tentadu")),
            attempted=Count("id", filter=Q(activity_result="Tentadu")),
        )
        context["stat_total"] = stats["total"]
        context["stat_achieved"] = stats["achieved"]
        context["stat_attempted"] = stats["attempted"]
        context["stat_rate"] = round(stats["achieved"] / stats["total"] * 100) if stats["total"] else 0

        return context


class TeacherActivityLogListView(LoginRequiredMixin, ListView):
    model = TeacherActivityLog
    template_name = "dashboards/teacher_logs.html"
    context_object_name = "logs"

    def get_queryset(self):
        from django.db.models import OuterRef, Subquery
        user = self.request.user

        daily_visits = (
            TeacherActivityLog.objects
            .exclude(sub_theme__isnull=True)
            .exclude(sub_theme='')
            .filter(
                teacher=OuterRef('teacher'),
                sub_theme=OuterRef('sub_theme'),
                activity_date=OuterRef('activity_date'),
            )
            .values('teacher')
            .annotate(cnt=Count('id'))
            .values('cnt')
        )

        queryset = (
            TeacherActivityLog.objects
            .select_related("preschool", "teacher")
            .exclude(sub_theme__isnull=True).exclude(sub_theme='')
            .annotate(vizita=Subquery(daily_visits))
            .order_by("-activity_date", "-created_at")
        )

        if user.role == "teacher":
            queryset = queryset.filter(teacher=user)
        elif user.role == "municipality_analyst":
            queryset = queryset.filter(preschool__municipality=user.municipality)
        elif user.role == "moe_admin":
            pass  # See all logs
        else:
            return queryset.none()

        # Apply filters
        teacher_id = self.request.GET.get('teacher')
        if teacher_id and user.role in ['moe_admin', 'municipality_analyst']:
            queryset = queryset.filter(teacher_id=teacher_id)

        preschool_id = self.request.GET.get('preschool')
        if preschool_id:
            queryset = queryset.filter(preschool_id=preschool_id)

        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)

        date_from = self.request.GET.get('date_from')
        if date_from:
            try:
                queryset = queryset.filter(created_at__gte=datetime.strptime(date_from, '%Y-%m-%d'))
            except ValueError:
                pass

        date_to = self.request.GET.get('date_to')
        if date_to:
            try:
                queryset = queryset.filter(created_at__lte=datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1))
            except ValueError:
                pass

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        context["page_title"] = "Teacher Logs"

        if user.role == "moe_admin":
            context["teachers"] = User.objects.filter(role='teacher').order_by('first_name')
            from preschools.models import Preschool
            context["preschools"] = Preschool.objects.all().order_by('name')
        elif user.role == "municipality_analyst":
            context["teachers"] = User.objects.filter(role='teacher', municipality=user.municipality).order_by('first_name')
            from preschools.models import Preschool
            context["preschools"] = Preschool.objects.filter(municipality=user.municipality).order_by('name')

        context["filter_teacher"] = self.request.GET.get('teacher', '')
        context["filter_preschool"] = self.request.GET.get('preschool', '')
        context["filter_status"] = self.request.GET.get('status', '')
        context["filter_date_from"] = self.request.GET.get('date_from', '')
        context["filter_date_to"] = self.request.GET.get('date_to', '')

        # Status summary: pivot per teacher → (theme, sub-theme) rows × date columns
        from collections import defaultdict
        summary_qs = (
            self.get_queryset()
            .exclude(sub_theme__isnull=True)
            .exclude(sub_theme='')
            .values(
                'teacher_id',
                'teacher__first_name',
                'teacher__last_name',
                'activity_date',
                'theme',
                'sub_theme',
            )
            .annotate(visits=Count('id'))
            .order_by('teacher__first_name', 'teacher__last_name', 'theme', 'sub_theme', '-activity_date')
        )

        pivot = {}
        for row in summary_qs:
            tid  = str(row['teacher_id'])
            name = f"{row['teacher__first_name']} {row['teacher__last_name']}"
            date = row['activity_date']
            key  = (row['theme'] or '', row['sub_theme'])
            if tid not in pivot:
                pivot[tid] = {'name': name, 'dates': set(), 'rows': defaultdict(dict),
                              'themes': {}}
            pivot[tid]['dates'].add(date)
            pivot[tid]['rows'][key][date] = row['visits']
            pivot[tid]['themes'][key] = row['theme'] or ''

        status_summary = []
        for tid, tdata in pivot.items():
            dates = sorted(tdata['dates'], reverse=True)
            sub_rows = []
            total_by_date = {d: 0 for d in dates}
            for key, day_counts in sorted(tdata['rows'].items()):
                cells = []
                row_total = 0
                for d in dates:
                    v = day_counts.get(d, 0)
                    cells.append(v)
                    row_total += v
                    total_by_date[d] += v
                sub_rows.append({
                    'theme': tdata['themes'][key],
                    'sub_theme': key[1],
                    'cells': cells,
                    'total': row_total,
                })
            status_summary.append({
                'name': tdata['name'],
                'dates': dates,
                'sub_rows': sub_rows,
                'totals': [total_by_date[d] for d in dates],
                'grand_total': sum(total_by_date.values()),
            })

        context['status_summary'] = status_summary
        context['active_tab'] = self.request.GET.get('tab', 'status')

        return context

    def render_to_response(self, context, **response_kwargs):
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest":
            html = render_to_string(
                "partials/teacher_logs_body.html",
                context,
                request=self.request
            )
            return JsonResponse({"html": html})
        return super().render_to_response(context, **response_kwargs)

    def get(self, request, *args, **kwargs):
        export_format = request.GET.get('export')
        if export_format == 'excel':
            return self.export_excel()
        elif export_format == 'pdf':
            return self.export_pdf()
        return super().get(request, *args, **kwargs)

    # ── shared pivot builder ───────────────────────────────────────────────
    def _build_pivot(self):
        from collections import defaultdict
        summary_qs = (
            self.get_queryset()
            .exclude(sub_theme__isnull=True).exclude(sub_theme='')
            .values('teacher_id','teacher__first_name','teacher__last_name',
                    'activity_date','theme','sub_theme')
            .annotate(visits=Count('id'))
            .order_by('teacher__first_name','teacher__last_name','theme','sub_theme','-activity_date')
        )
        pivot = {}
        for row in summary_qs:
            tid  = str(row['teacher_id'])
            name = f"{row['teacher__first_name']} {row['teacher__last_name']}"
            date = row['activity_date']
            key  = (row['theme'] or '', row['sub_theme'])
            if tid not in pivot:
                pivot[tid] = {'name': name, 'dates': set(), 'rows': defaultdict(dict),
                              'themes': {}}
            pivot[tid]['dates'].add(date)
            pivot[tid]['rows'][key][date] = row['visits']
            pivot[tid]['themes'][key] = row['theme'] or ''

        result = []
        for tdata in pivot.values():
            dates = sorted(tdata['dates'], reverse=True)
            sub_rows, total_by_date = [], {d: 0 for d in dates}
            for key, day_counts in sorted(tdata['rows'].items()):
                cells, row_total = [], 0
                for d in dates:
                    v = day_counts.get(d, 0)
                    cells.append(v); row_total += v; total_by_date[d] += v
                sub_rows.append({
                    'theme': tdata['themes'][key],
                    'sub_theme': key[1],
                    'cells': cells,
                    'total': row_total,
                })
            result.append({
                'name': tdata['name'], 'dates': dates, 'sub_rows': sub_rows,
                'totals': [total_by_date[d] for d in dates],
                'grand_total': sum(total_by_date.values()),
            })
        return result

    # ── shared export helpers ──────────────────────────────────────────────
    def _export_meta(self):
        """Return (teacher_label, period_label, queryset) for export headers."""
        qs = self.get_queryset()
        teacher_id = self.request.GET.get('teacher', '')
        date_from  = self.request.GET.get('date_from', '')
        date_to    = self.request.GET.get('date_to', '')

        if teacher_id:
            try:
                t = User.objects.get(pk=teacher_id)
                teacher_label = t.get_full_name() or t.username
            except User.DoesNotExist:
                teacher_label = 'Profesór Hotu'
        else:
            teacher_label = 'Profesór Hotu'

        parts = []
        if date_from: parts.append(f"Husi {date_from}")
        if date_to:   parts.append(f"To'o {date_to}")
        period_label = '  |  '.join(parts) if parts else 'Períodu Hotu'

        return teacher_label, period_label, qs

    # ── Excel export ───────────────────────────────────────────────────────
    def export_excel(self):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import date as dt_date

        teacher_label, period_label, qs = self._export_meta()
        records = list(qs)
        wb = openpyxl.Workbook()

        BLUE_DARK = "1D4ED8"
        BLUE_MID  = "2563EB"
        BLUE_PALE = "EFF6FF"
        BLUE_LITE = "DBEAFE"
        WHITE     = "FFFFFF"
        GREY_ROW  = "F8FAFC"
        GREY_LINE = "E2E8F0"
        SLATE     = "334155"

        thin = Side(style='thin', color=GREY_LINE)
        bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

        def _c(ws, r, col, val, bold=False, fg=None, bg=None,
               align='left', size=10):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font      = Font(name='Calibri', bold=bold,
                                  color=fg or SLATE, size=size)
            cell.alignment = Alignment(horizontal=align, vertical='center')
            if bg:
                cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = bdr
            return cell

        ws = wb.active
        ws.title = "Vizita Profesór"
        ws.sheet_view.showGridLines = False

        COL_WIDTHS = [6, 14, 22, 22, 18, 22, 10]
        HEADERS    = ["#", "Data", "Profesór", "Pre-Eskolár", "Tema", "Tipo", "Vizita"]
        N = len(HEADERS)

        # ── Cover rows ─────────────────────────────────────────────────────
        r = 1
        ws.row_dimensions[r].height = 44
        c = ws.cell(row=r, column=1, value="HAAP — Relatóriu Vizita Profesór")
        c.font      = Font(name='Calibri', bold=True, size=20, color=WHITE)
        c.fill      = PatternFill("solid", fgColor=BLUE_MID)
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N)
        r += 1

        for label, val in [
            ("Profesór",  teacher_label),
            ("Período",   period_label),
            ("Gerado em", dt_date.today().strftime('%d %B %Y')),
        ]:
            ws.row_dimensions[r].height = 18
            c = ws.cell(row=r, column=1, value=f"{label}: {val}")
            c.font      = Font(name='Calibri', size=10, color="64748B")
            c.fill      = PatternFill("solid", fgColor=BLUE_PALE)
            c.alignment = Alignment(horizontal='left', vertical='center')
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N)
            r += 1
        r += 1  # blank row

        # ── Column headers ─────────────────────────────────────────────────
        ws.row_dimensions[r].height = 26
        for ci, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
            c = ws.cell(row=r, column=ci, value=h)
            c.font      = Font(name='Calibri', bold=True, size=10, color=WHITE)
            c.fill      = PatternFill("solid", fgColor=BLUE_DARK)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border    = bdr
            ws.column_dimensions[get_column_letter(ci)].width = w
        r += 1

        # ── Data rows ──────────────────────────────────────────────────────
        for idx, log in enumerate(records, start=1):
            bg = WHITE if idx % 2 == 0 else GREY_ROW
            ws.row_dimensions[r].height = 18
            date_val = log.activity_date
            if hasattr(date_val, 'date'):
                date_val = date_val.date()
            row_data = [
                idx,
                date_val.strftime('%d/%m/%Y') if date_val else '—',
                log.teacher.get_full_name() or log.teacher.username,
                log.preschool.name if log.preschool else '—',
                log.theme or '—',
                log.sub_theme or '—',
                log.status,
            ]
            center_cols = {1, 2, 7}
            for ci, val in enumerate(row_data, start=1):
                c = ws.cell(row=r, column=ci, value=val)
                c.font      = Font(name='Calibri', size=9, color=SLATE)
                c.fill      = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(
                    horizontal='center' if ci in center_cols else 'left',
                    vertical='center')
                c.border = bdr
            r += 1

        # ── Summary row ────────────────────────────────────────────────────
        ws.row_dimensions[r].height = 22
        c = ws.cell(row=r, column=1, value=f"Total: {len(records)} rejistu")
        c.font      = Font(name='Calibri', bold=True, size=10, color=WHITE)
        c.fill      = PatternFill("solid", fgColor=BLUE_DARK)
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        fname = teacher_label.replace(' ', '_')
        response['Content-Disposition'] = f'attachment; filename="haap_vizita_{fname}.xlsx"'
        return response

    # ── PDF export ─────────────────────────────────────────────────────────
    def export_pdf(self):
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from datetime import date as dt_date

        teacher_label, period_label, qs = self._export_meta()
        records = list(qs)

        buffer = BytesIO()
        MARGIN = 1.5 * cm
        doc    = SimpleDocTemplate(
            buffer, pagesize=A4,
            leftMargin=MARGIN, rightMargin=MARGIN,
            topMargin=MARGIN, bottomMargin=MARGIN,
        )
        PW = A4[0] - 2 * MARGIN

        BLUE      = colors.HexColor('#2563EB')
        BLUE_DARK = colors.HexColor('#1D4ED8')
        BLUE_PALE = colors.HexColor('#EFF6FF')
        BLUE_LITE = colors.HexColor('#DBEAFE')
        GREY_ROW  = colors.HexColor('#F8FAFC')
        GREY_LINE = colors.HexColor('#E2E8F0')
        SLATE     = colors.HexColor('#334155')
        WHITE     = colors.white
        ACCENT    = colors.HexColor('#93C5FD')

        styles = getSampleStyleSheet()
        s_normal = ParagraphStyle('p_n', parent=styles['Normal'],
                                  fontSize=8, fontName='Helvetica',
                                  textColor=SLATE, leading=11)
        s_bold   = ParagraphStyle('p_b', parent=styles['Normal'],
                                  fontSize=8, fontName='Helvetica-Bold',
                                  textColor=SLATE, leading=11)
        s_hdr    = ParagraphStyle('p_h', parent=styles['Normal'],
                                  fontSize=8, fontName='Helvetica-Bold',
                                  textColor=WHITE, alignment=TA_CENTER, leading=11)
        s_center = ParagraphStyle('p_c', parent=styles['Normal'],
                                  fontSize=8, fontName='Helvetica',
                                  textColor=SLATE, alignment=TA_CENTER, leading=11)

        elements = []

        # ── Page header ────────────────────────────────────────────────────
        hdr_data = [[
            Paragraph("<b>HAAP</b><br/><font size='9' color='#93C5FD'>Relatóriu Vizita Profesór</font>",
                      ParagraphStyle('ht', parent=styles['Normal'], fontSize=18,
                                     fontName='Helvetica-Bold', textColor=WHITE,
                                     leading=24)),
            Paragraph(
                f"<b>{teacher_label}</b><br/>"
                f"<font size='8' color='#93C5FD'>{period_label}</font><br/>"
                f"<font size='7' color='#DBEAFE'>Gerado: {dt_date.today().strftime('%d %B %Y')}</font>",
                ParagraphStyle('hm', parent=styles['Normal'], fontSize=11,
                               fontName='Helvetica-Bold', textColor=WHITE,
                               alignment=TA_RIGHT, leading=16)),
        ]]
        hdr_tbl = Table(hdr_data, colWidths=[PW * 0.55, PW * 0.45])
        hdr_tbl.setStyle(TableStyle([
            ('BACKGROUND',    (0,0), (-1,-1), BLUE),
            ('TOPPADDING',    (0,0), (-1,-1), 14),
            ('BOTTOMPADDING', (0,0), (-1,-1), 14),
            ('LEFTPADDING',   (0,0), (0,0),   16),
            ('RIGHTPADDING',  (1,0), (1,0),   16),
            ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
            ('LINEBELOW',     (0,0), (-1,-1), 3, BLUE_DARK),
        ]))
        elements += [hdr_tbl, Spacer(1, 0.5*cm)]

        # ── Data table ─────────────────────────────────────────────────────
        col_w = [0.7*cm, 2.2*cm, 3.8*cm, 3.2*cm, 3.2*cm, 4.0*cm, 1.4*cm]
        hdr_row = [Paragraph(h, s_hdr)
                   for h in ["#", "Data", "Profesór", "Pre-Eskolár", "Tema", "Tipo", "Vizita"]]
        tbl_data = [hdr_row]
        ts = [
            ('BACKGROUND',    (0,0), (-1,0), BLUE),
            ('FONTNAME',      (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',      (0,0), (-1,-1), 8),
            ('TOPPADDING',    (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('GRID',          (0,0), (-1,-1), 0.4, GREY_LINE),
            ('ALIGN',         (0,0), (-1,-1), 'LEFT'),
            ('ALIGN',         (0,0), (0,-1), 'CENTER'),
            ('ALIGN',         (1,0), (1,-1), 'CENTER'),
            ('ALIGN',         (6,0), (6,-1), 'CENTER'),
            ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ]

        for idx, log in enumerate(records, start=1):
            bg = WHITE if idx % 2 == 0 else GREY_ROW
            date_val = log.activity_date
            if hasattr(date_val, 'date'):
                date_val = date_val.date()
            row = [
                Paragraph(str(idx), s_center),
                Paragraph(date_val.strftime('%d/%m/%Y') if date_val else '—', s_center),
                Paragraph(log.teacher.get_full_name() or log.teacher.username, s_normal),
                Paragraph(log.preschool.name if log.preschool else '—', s_normal),
                Paragraph(log.theme or '—', s_normal),
                Paragraph(log.sub_theme or '—', s_bold),
                Paragraph(str(log.status), s_center),
            ]
            tbl_data.append(row)
            ts.append(('BACKGROUND', (0, idx), (-1, idx), bg))

        # Total row
        total_row = [
            Paragraph(f"<b>Total: {len(records)}</b>", s_bold),
            *[''] * 6,
        ]
        tbl_data.append(total_row)
        last = len(tbl_data) - 1
        ts += [
            ('BACKGROUND', (0, last), (-1, last), BLUE_LITE),
            ('FONTNAME',   (0, last), (-1, last), 'Helvetica-Bold'),
            ('SPAN',       (0, last), (-1, last)),
        ]

        tbl = Table(tbl_data, colWidths=col_w, repeatRows=1)
        tbl.setStyle(TableStyle(ts))
        elements.append(tbl)

        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        fname = teacher_label.replace(' ', '_')
        response['Content-Disposition'] = f'attachment; filename="haap_vizita_{fname}.pdf"'
        return response
