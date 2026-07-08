"""
System Audit Log views — accessible to moe_admin only.
"""
import csv
import io
from datetime import datetime, timedelta

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Count, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from core.models import AuditLog, User
from core.audit import log_action


def _require_admin(request):
    """Return redirect response if not moe_admin, else None."""
    if not request.user.is_authenticated or request.user.role != 'moe_admin':
        return redirect('core:login')
    return None


def _apply_filters(request, qs):
    search    = request.GET.get('q', '').strip()
    user_id   = request.GET.get('user', '').strip()
    action    = request.GET.get('action', '').strip()
    module    = request.GET.get('module', '').strip()
    status    = request.GET.get('status', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to   = request.GET.get('date_to', '').strip()

    if search:
        qs = qs.filter(
            Q(username__icontains=search) |
            Q(description__icontains=search) |
            Q(record_name__icontains=search) |
            Q(ip_address__icontains=search) |
            Q(module__icontains=search)
        )
    if user_id:
        qs = qs.filter(user_id=user_id)
    if action:
        qs = qs.filter(action=action)
    if module:
        qs = qs.filter(module__iexact=module)
    if status:
        qs = qs.filter(status=status)
    if date_from:
        try:
            qs = qs.filter(timestamp__date__gte=date_from)
        except Exception:
            pass
    if date_to:
        try:
            qs = qs.filter(timestamp__date__lte=date_to)
        except Exception:
            pass
    return qs


@login_required
def audit_list(request):
    guard = _require_admin(request)
    if guard:
        return guard

    base_qs = AuditLog.objects.select_related('user')

    # ── Sorting ──────────────────────────────────────────────────────────────
    sort = request.GET.get('sort', '-timestamp')
    valid_sorts = {
        'timestamp': 'timestamp', '-timestamp': '-timestamp',
        'username': 'username',   '-username': '-username',
        'action': 'action',       '-action': '-action',
        'module': 'module',       '-module': '-module',
        'status': 'status',       '-status': '-status',
    }
    qs = _apply_filters(request, base_qs).order_by(valid_sorts.get(sort, '-timestamp'))

    # ── Stats ─────────────────────────────────────────────────────────────────
    now   = timezone.now()
    today = now.date()
    week  = today - timedelta(days=6)

    total_all     = AuditLog.objects.count()
    today_count   = AuditLog.objects.filter(timestamp__date=today).count()
    week_count    = AuditLog.objects.filter(timestamp__date__gte=week).count()
    failed_week   = AuditLog.objects.filter(status='failed', timestamp__date__gte=week).count()

    # Daily activity (last 7 days) for chart
    daily = []
    for i in range(6, -1, -1):
        d = today - timedelta(days=i)
        daily.append({
            'date': d.strftime('%d/%m'),
            'count': AuditLog.objects.filter(timestamp__date=d).count(),
        })

    # Top users this week
    top_users = (
        AuditLog.objects
        .filter(timestamp__date__gte=week)
        .exclude(username='')
        .values('username')
        .annotate(cnt=Count('id'))
        .order_by('-cnt')[:5]
    )

    # ── Pagination ────────────────────────────────────────────────────────────
    paginator = Paginator(qs, 25)
    page_obj  = paginator.get_page(request.GET.get('page', 1))

    # ── Filter UI helpers ─────────────────────────────────────────────────────
    users   = (
        User.objects
        .filter(audit_logs__isnull=False)
        .distinct()
        .order_by('first_name', 'last_name')
    )
    modules = (
        AuditLog.objects
        .exclude(module='')
        .values_list('module', flat=True)
        .distinct()
        .order_by('module')
    )

    log_action(
        request=request,
        action='view', module='Audit',
        description='Haree sistema audit log.',
    )

    return render(request, 'audit/audit_list.html', {
        'page_obj':       page_obj,
        'total_all':      total_all,
        'today_count':    today_count,
        'week_count':     week_count,
        'failed_week':    failed_week,
        'daily':          daily,
        'top_users':      top_users,
        'users':          users,
        'modules':        modules,
        'action_choices': AuditLog.ACTION_CHOICES,
        'filters': {
            'q':         request.GET.get('q', ''),
            'user':      request.GET.get('user', ''),
            'action':    request.GET.get('action', ''),
            'module':    request.GET.get('module', ''),
            'status':    request.GET.get('status', ''),
            'date_from': request.GET.get('date_from', ''),
            'date_to':   request.GET.get('date_to', ''),
            'sort':      sort,
        },
    })


@login_required
def audit_detail(request, log_id):
    guard = _require_admin(request)
    if guard:
        return JsonResponse({'error': 'Aksesu negadu.'}, status=403)
    log = get_object_or_404(AuditLog, id=log_id)
    return JsonResponse({
        'id':             str(log.id),
        'username':       log.username,
        'action':         log.get_action_display(),
        'action_raw':     log.action,
        'module':         log.module,
        'description':    log.description,
        'record_id':      log.record_id,
        'record_name':    log.record_name,
        'previous_value': log.previous_value,
        'new_value':      log.new_value,
        'ip_address':     log.ip_address or '—',
        'browser':        log.browser or '—',
        'os_info':        log.os_info or '—',
        'user_agent':     log.user_agent,
        'status':         log.get_status_display(),
        'status_raw':     log.status,
        'timestamp':      log.timestamp.strftime('%d/%m/%Y %H:%M:%S'),
    })


@login_required
def audit_export(request):
    guard = _require_admin(request)
    if guard:
        return HttpResponse('Aksesu negadu.', status=403)

    fmt = request.GET.get('format', 'csv')
    qs  = _apply_filters(
        request,
        AuditLog.objects.select_related('user').order_by('-timestamp'),
    )

    log_action(
        request=request,
        action='export', module='Audit',
        description=f'Export audit log — formato: {fmt.upper()}',
    )

    if fmt == 'excel':
        return _export_excel(qs)
    elif fmt == 'pdf':
        return _export_pdf(qs)
    return _export_csv(qs)


# ── Export helpers ────────────────────────────────────────────────────────────

_HEADERS = [
    'Timestamp', 'Utilizador', 'Aksaun', 'Modulu', 'Deskrisaun',
    'Rekord ID', 'Rekord Naran', 'IP Address', 'Browser', 'OS', 'Status',
]


def _row(log):
    return [
        log.timestamp.strftime('%d/%m/%Y %H:%M:%S'),
        log.username,
        log.get_action_display(),
        log.module,
        log.description,
        log.record_id,
        log.record_name,
        log.ip_address or '',
        log.browser,
        log.os_info,
        log.get_status_display(),
    ]


def _export_csv(qs):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="audit_log.csv"'
    response.write('﻿')
    writer = csv.writer(response)
    writer.writerow(_HEADERS)
    for log in qs[:10_000]:
        writer.writerow(_row(log))
    return response


def _export_excel(qs):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = 'Audit Log'

    hdr_font  = Font(bold=True, color='FFFFFF', size=10)
    hdr_fill  = PatternFill(start_color='1D4ED8', end_color='1D4ED8', fill_type='solid')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin      = Side(style='thin', color='CBD5E1')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(_HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, hdr_align, border

    alt_fill = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
    for r, log in enumerate(qs[:10_000], 2):
        for col, val in enumerate(_row(log), 1):
            c = ws.cell(row=r, column=col, value=val)
            c.border = border
            if r % 2 == 0:
                c.fill = alt_fill

    col_widths = [20, 22, 16, 14, 50, 14, 22, 16, 18, 14, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = 'A2'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="audit_log.xlsx"'
    wb.save(response)
    return response


def _export_pdf(qs):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="audit_log.pdf"'

    doc = SimpleDocTemplate(
        response, pagesize=landscape(A4),
        topMargin=1 * cm, bottomMargin=1 * cm,
        leftMargin=1 * cm, rightMargin=1 * cm,
    )
    styles = getSampleStyleSheet()
    elements = []
    elements.append(Paragraph('HAAP — Sistema Audit Log', styles['Title']))
    elements.append(Paragraph(
        f'Jeradu: {datetime.now().strftime("%d/%m/%Y %H:%M")}',
        styles['Normal'],
    ))
    elements.append(Spacer(1, 0.4 * cm))

    pdf_headers = ['Timestamp', 'Utilizador', 'Aksaun', 'Modulu', 'Deskrisaun', 'IP', 'Status']
    data = [pdf_headers]
    for log in qs[:5_000]:
        desc = log.description
        if len(desc) > 60:
            desc = desc[:57] + '...'
        data.append([
            log.timestamp.strftime('%d/%m/%Y\n%H:%M'),
            log.username[:18],
            log.get_action_display(),
            log.module[:12],
            desc,
            log.ip_address or '—',
            log.get_status_display(),
        ])

    col_w = [3.0*cm, 3.5*cm, 2.5*cm, 2.5*cm, 10*cm, 3.0*cm, 2.5*cm]
    t = Table(data, colWidths=col_w, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0),  colors.HexColor('#1D4ED8')),
        ('TEXTCOLOR',     (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',      (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, 0),  8),
        ('FONTSIZE',      (0, 1), (-1, -1), 7),
        ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, colors.HexColor('#EFF6FF')]),
        ('GRID',          (0, 0), (-1, -1), 0.3, colors.HexColor('#CBD5E1')),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',    (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(t)
    doc.build(elements)
    return response
