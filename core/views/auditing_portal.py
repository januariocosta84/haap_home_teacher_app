"""
Auditing Portal views — accessible exclusively to moe_auditing role.

Routes live under /auditing/ and use separate templates so the UI is
completely isolated from the moe_admin dashboard.
"""
import csv
from datetime import datetime, timedelta
from functools import wraps

from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.core.paginator import Paginator
from django.db.models import Count, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST

from core.models import AuditLog, User
from core.audit import log_action


# ── Access guard ──────────────────────────────────────────────────────────────

def require_auditing(view_fn):
    """Decorator: allow only authenticated moe_auditing users."""
    @wraps(view_fn)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('core:auditing_login')
        if request.user.role != 'moe_auditing':
            return redirect('core:auditing_login')
        return view_fn(request, *args, **kwargs)
    return wrapper


# ── Authentication ────────────────────────────────────────────────────────────

def auditing_login(request):
    """Email + password login exclusively for moe_auditing users."""
    if request.user.is_authenticated and request.user.role == 'moe_auditing':
        return redirect('core:auditing_dashboard')

    error = None
    if request.method == 'POST':
        email    = request.POST.get('email', '').strip()
        password = request.POST.get('password', '').strip()

        if not email or not password:
            error = 'Favor hakerek email no password.'
        else:
            user = authenticate(request, email=email, password=password)
            if user is not None and user.role == 'moe_auditing':
                login(request, user, backend='core.backends.EmailAuditingBackend')
                log_action(
                    request=request, user=user,
                    action='login', module='Auditing Portal',
                    description=f'{user.get_full_name()} login ba auditing portal.',
                )
                return redirect('core:auditing_dashboard')
            else:
                log_action(
                    request=request, user=None,
                    action='login_failed', module='Auditing Portal',
                    description=f'Login falha ba email: {email}',
                    username=email,
                    status='failed',
                )
                error = 'Email ka password sala. Ka kontu ne\'e la iha permisaun auditing.'

    return render(request, 'auditing/login.html', {'error': error})


def auditing_logout(request):
    if request.user.is_authenticated and request.user.role == 'moe_auditing':
        log_action(
            request=request,
            action='logout', module='Auditing Portal',
            description=f'{request.user.get_full_name()} logout husi auditing portal.',
        )
    logout(request)
    return redirect('core:auditing_login')


# ── Dashboard ─────────────────────────────────────────────────────────────────

@require_auditing
def auditing_dashboard(request):
    now   = timezone.now()
    today = now.date()
    week  = today - timedelta(days=6)
    month = today - timedelta(days=29)

    # Stat cards
    total_logs    = AuditLog.objects.count()
    today_logs    = AuditLog.objects.filter(timestamp__date=today).count()
    failed_week   = AuditLog.objects.filter(status='failed', timestamp__date__gte=week).count()
    active_users  = (
        AuditLog.objects
        .filter(timestamp__date__gte=week)
        .exclude(username='')
        .values('username').distinct().count()
    )

    # Daily breakdown — last 14 days
    daily = []
    for i in range(13, -1, -1):
        d = today - timedelta(days=i)
        cnt = AuditLog.objects.filter(timestamp__date=d).count()
        daily.append({'date': d.strftime('%d/%m'), 'count': cnt})

    # Action distribution (last 30 days)
    action_dist = (
        AuditLog.objects
        .filter(timestamp__date__gte=month)
        .values('action')
        .annotate(cnt=Count('id'))
        .order_by('-cnt')[:10]
    )
    action_labels = [d['action'] for d in action_dist]
    action_counts = [d['cnt']   for d in action_dist]

    # Recent security events (failed logins last 7 days)
    security_events = (
        AuditLog.objects
        .filter(status='failed', timestamp__date__gte=week)
        .order_by('-timestamp')[:10]
    )

    # Recent login/logout
    auth_logs = (
        AuditLog.objects
        .filter(action__in=['login', 'logout', 'login_failed'])
        .order_by('-timestamp')[:15]
    )

    # Top active users (week)
    top_users = (
        AuditLog.objects
        .filter(timestamp__date__gte=week)
        .exclude(username='')
        .values('username')
        .annotate(cnt=Count('id'))
        .order_by('-cnt')[:8]
    )

    # Recent activity feed
    recent = AuditLog.objects.order_by('-timestamp')[:20]

    return render(request, 'auditing/dashboard.html', {
        'total_logs':     total_logs,
        'today_logs':     today_logs,
        'failed_week':    failed_week,
        'active_users':   active_users,
        'daily':          daily,
        'action_labels':  action_labels,
        'action_counts':  action_counts,
        'security_events': security_events,
        'auth_logs':      auth_logs,
        'top_users':      top_users,
        'recent':         recent,
    })


# ── Full log list ─────────────────────────────────────────────────────────────

def _apply_log_filters(request, qs):
    search    = request.GET.get('q', '').strip()
    user_id   = request.GET.get('user', '').strip()
    action    = request.GET.get('action', '').strip()
    module    = request.GET.get('module', '').strip()
    status    = request.GET.get('status', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to   = request.GET.get('date_to', '').strip()

    if search:
        qs = qs.filter(
            Q(username__icontains=search) |
            Q(description__icontains=search) |
            Q(record_name__icontains=search) |
            Q(ip_address__icontains=search) |
            Q(module__icontains=search)
        )
    if user_id:
        qs = qs.filter(user_id=user_id)
    if action:
        qs = qs.filter(action=action)
    if module:
        qs = qs.filter(module__iexact=module)
    if status:
        qs = qs.filter(status=status)
    if date_from:
        try:
            qs = qs.filter(timestamp__date__gte=date_from)
        except Exception:
            pass
    if date_to:
        try:
            qs = qs.filter(timestamp__date__lte=date_to)
        except Exception:
            pass
    return qs


@require_auditing
def auditing_logs(request):
    sort = request.GET.get('sort', '-timestamp')
    valid_sorts = {
        'timestamp': 'timestamp',   '-timestamp': '-timestamp',
        'username':  'username',    '-username':  '-username',
        'action':    'action',      '-action':    '-action',
        'module':    'module',      '-module':    '-module',
        'status':    'status',      '-status':    '-status',
    }
    qs = _apply_log_filters(
        request,
        AuditLog.objects.select_related('user'),
    ).order_by(valid_sorts.get(sort, '-timestamp'))

    paginator = Paginator(qs, 30)
    page_obj  = paginator.get_page(request.GET.get('page', 1))

    users   = User.objects.filter(audit_logs__isnull=False).distinct().order_by('first_name')
    modules = (
        AuditLog.objects.exclude(module='')
        .values_list('module', flat=True).distinct().order_by('module')
    )

    return render(request, 'auditing/logs.html', {
        'page_obj':       page_obj,
        'users':          users,
        'modules':        modules,
        'action_choices': AuditLog.ACTION_CHOICES,
        'filters': {
            'q':         request.GET.get('q', ''),
            'user':      request.GET.get('user', ''),
            'action':    request.GET.get('action', ''),
            'module':    request.GET.get('module', ''),
            'status':    request.GET.get('status', ''),
            'date_from': request.GET.get('date_from', ''),
            'date_to':   request.GET.get('date_to', ''),
            'sort':      sort,
        },
        'total': qs.count(),
    })


@require_auditing
def auditing_log_detail(request, log_id):
    log = get_object_or_404(AuditLog, id=log_id)
    return JsonResponse({
        'id':             str(log.id),
        'username':       log.username,
        'action':         log.get_action_display(),
        'action_raw':     log.action,
        'module':         log.module,
        'description':    log.description,
        'record_id':      log.record_id,
        'record_name':    log.record_name,
        'previous_value': log.previous_value,
        'new_value':      log.new_value,
        'ip_address':     log.ip_address or '—',
        'browser':        log.browser or '—',
        'os_info':        log.os_info or '—',
        'user_agent':     log.user_agent,
        'status':         log.get_status_display(),
        'status_raw':     log.status,
        'timestamp':      log.timestamp.strftime('%d/%m/%Y %H:%M:%S'),
    })


# ── Export ────────────────────────────────────────────────────────────────────

_HEADERS = [
    'Timestamp', 'Utilizador', 'Aksaun', 'Modulu', 'Deskrisaun',
    'Rekord ID', 'Rekord Naran', 'IP Address', 'Browser', 'OS', 'Status',
]


def _row(log):
    return [
        log.timestamp.strftime('%d/%m/%Y %H:%M:%S'),
        log.username, log.get_action_display(), log.module,
        log.description, log.record_id, log.record_name,
        log.ip_address or '', log.browser, log.os_info,
        log.get_status_display(),
    ]


@require_auditing
def auditing_export(request):
    fmt = request.GET.get('format', 'csv')
    qs  = _apply_log_filters(
        request,
        AuditLog.objects.select_related('user').order_by('-timestamp'),
    )
    log_action(
        request=request,
        action='export', module='Auditing Portal',
        description=f'Export audit log — formato: {fmt.upper()}',
    )
    if fmt == 'excel':
        return _export_excel(qs)
    elif fmt == 'pdf':
        return _export_pdf(qs)
    return _export_csv(qs)


def _export_csv(qs):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="audit_log.csv"'
    response.write('﻿')
    writer = csv.writer(response)
    writer.writerow(_HEADERS)
    for log in qs[:10_000]:
        writer.writerow(_row(log))
    return response


def _export_excel(qs):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = 'Audit Log'

    hdr_font  = Font(bold=True, color='FFFFFF', size=10)
    hdr_fill  = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin      = Side(style='thin', color='CBD5E1')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill  = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')

    for col, h in enumerate(_HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, hdr_align, border

    for r, log in enumerate(qs[:10_000], 2):
        for col, val in enumerate(_row(log), 1):
            c = ws.cell(row=r, column=col, value=val)
            c.border = border
            if r % 2 == 0:
                c.fill = alt_fill

    for i, w in enumerate([20, 22, 16, 14, 50, 14, 22, 16, 18, 14, 12], 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = 'A2'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="audit_log.xlsx"'
    wb.save(response)
    return response


def _export_pdf(qs):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="audit_log.pdf"'

    doc = SimpleDocTemplate(
        response, pagesize=landscape(A4),
        topMargin=1*cm, bottomMargin=1*cm, leftMargin=1*cm, rightMargin=1*cm,
    )
    styles  = getSampleStyleSheet()
    elements = [
        Paragraph('HAAP — Sistema Audit Log', styles['Title']),
        Paragraph(f'Jeradu: {datetime.now().strftime("%d/%m/%Y %H:%M")}', styles['Normal']),
        Spacer(1, 0.4*cm),
    ]

    pdf_headers = ['Timestamp', 'Utilizador', 'Aksaun', 'Modulu', 'Deskrisaun', 'IP', 'Status']
    data = [pdf_headers]
    for log in qs[:5_000]:
        desc = log.description
        if len(desc) > 60:
            desc = desc[:57] + '...'
        data.append([
            log.timestamp.strftime('%d/%m/%Y\n%H:%M'),
            log.username[:18], log.get_action_display(),
            log.module[:12], desc, log.ip_address or '—',
            log.get_status_display(),
        ])

    t = Table(data, colWidths=[3*cm, 3.5*cm, 2.5*cm, 2.5*cm, 10*cm, 3*cm, 2.5*cm], repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND',    (0,0),(-1,0),  colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR',     (0,0),(-1,0),  colors.white),
        ('FONTNAME',      (0,0),(-1,0),  'Helvetica-Bold'),
        ('FONTSIZE',      (0,0),(-1,0),  8),
        ('FONTSIZE',      (0,1),(-1,-1), 7),
        ('ROWBACKGROUNDS',(0,1),(-1,-1), [colors.white, colors.HexColor('#EFF6FF')]),
        ('GRID',          (0,0),(-1,-1), 0.3, colors.HexColor('#CBD5E1')),
        ('VALIGN',        (0,0),(-1,-1), 'MIDDLE'),
        ('TOPPADDING',    (0,0),(-1,-1), 3),
        ('BOTTOMPADDING', (0,0),(-1,-1), 3),
    ]))
    elements.append(t)
    doc.build(elements)
    return response
